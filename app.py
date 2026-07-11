from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import sqlite3
import pandas as pd
import os
import io
import sys
import socket
import json
import logging
import shutil
import secrets
import re
import urllib.request
import urllib.error
import qrcode
from PIL import Image, UnidentifiedImageError

# Evita UnicodeEncodeError al imprimir emojis en consolas Windows (cp1252)
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, 'reconfigure'):
        _stream.reconfigure(encoding='utf-8', errors='replace')
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, 'latas.db')
EXCEL_PATH = os.path.join(BASE_DIR, 'Latas de cerveza.xlsx')
FOTOS_DIR  = os.path.join(BASE_DIR, 'static', 'fotos')
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')

# El token de sincronización con el buzón en la nube es una credencial, así que
# vive un nivel arriba de la carpeta del proyecto (la "carpeta madre", ej. el
# Escritorio) en vez de adentro: así queda físicamente fuera del repo git —
# nunca puede subirse a GitHub pase lo que pase con el .gitignore — pero sigue
# siendo una carpeta normal que se ve en el Explorador de Windows.
LOCAL_CONFIG_DIR = os.path.join(os.path.dirname(BASE_DIR), 'LatasPepe-config')
SYNC_CONFIG_PATH = os.path.join(LOCAL_CONFIG_DIR, 'sync_config.json')

# Este mismo app.py se despliega dos veces: acá en la PC (fuente de verdad,
# lectura y escritura completa) y en PythonAnywhere (app.py del deploy tiene
# al lado un cloud_config.json que NUNCA existe en la PC ni se sube a GitHub
# — vive solo en el servidor). Su sola presencia activa el modo nube: la
# colección se ve completa (dashboard, galería, wishlist) pero de solo
# lectura; lo único que se puede hacer ahí es anotar una lata nueva, que
# queda pendiente hasta que la PC sincroniza.
CLOUD_CONFIG_PATH = os.path.join(BASE_DIR, 'cloud_config.json')
CLOUD_MODE = os.path.exists(CLOUD_CONFIG_PATH)

MAX_PER_PAGE = 200
FOTO_EXTS    = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
PORT         = 5000

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 12 * 1024 * 1024  # 12 MB por foto

CLOUD_PIN = None
if CLOUD_MODE:
    with open(CLOUD_CONFIG_PATH, encoding='utf-8') as f:
        _cloud_cfg = json.load(f)
    app.secret_key = _cloud_cfg['secret_key']
    CLOUD_PIN = str(_cloud_cfg['pin'])

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = app.logger

# ── Mappings ──────────────────────────────────────────────────────────────────

PAIS_ISO = {
    'Alemania': 'DE', 'Angola': 'AO', 'Argentina': 'AR', 'Austria': 'AT',
    'Bélgica': 'BE', 'Bolivia': 'BO', 'Brasil': 'BR', 'Cabo Verde': 'CV',
    'Canadá': 'CA', 'Chile': 'CL', 'China': 'CN', 'Colombia': 'CO',
    'Corea del Sur': 'KR', 'Costa Rica': 'CR', 'Croacia': 'HR',
    'Dinamarca': 'DK', 'E.E.U.U.': 'US', 'Escocia': 'GB', 'Eslovaquia': 'SK',
    'Eslovenia': 'SI', 'España': 'ES', 'Filipinas': 'PH', 'Finlandia': 'FI',
    'Francia': 'FR', 'Grecia': 'GR', 'Holanda': 'NL', 'Hungría': 'HU',
    'India': 'IN', 'Inglaterra': 'GB', 'Irlanda': 'IE', 'Italia': 'IT',
    'Italia (Cerdeña)': 'IT', 'Japón': 'JP', 'Lituania': 'LT', 'México': 'MX',
    'Nicaragua': 'NI', 'Nueva Zelanda': 'NZ', 'Panamá': 'PA', 'Paraguay': 'PY',
    'Perú': 'PE', 'Polonia': 'PL', 'Portugal': 'PT', 'Rep. Checa': 'CZ',
    'Rep. Dominicana': 'DO', 'Rumania': 'RO', 'Serbia': 'RS', 'Sudáfrica': 'ZA',
    'Suiza': 'CH', 'Tailandia': 'TH', 'Turquía': 'TR', 'Uruguay': 'UY',
    'Venezuela': 'VE',
}

CONTINENTES = {
    'Europa': [
        'Alemania', 'Austria', 'Bélgica', 'Croacia', 'Dinamarca', 'Escocia',
        'Eslovaquia', 'Eslovenia', 'España', 'Finlandia', 'Francia', 'Grecia',
        'Holanda', 'Hungría', 'Inglaterra', 'Irlanda', 'Italia', 'Italia (Cerdeña)',
        'Lituania', 'Polonia', 'Portugal', 'Rep. Checa', 'Rumania', 'Serbia', 'Suiza', 'Turquía',
    ],
    'América del Sur': [
        'Argentina', 'Bolivia', 'Brasil', 'Chile', 'Colombia',
        'Paraguay', 'Perú', 'Uruguay', 'Venezuela',
    ],
    'América del Norte': ['Canadá', 'E.E.U.U.', 'México'],
    'C. América y Caribe': ['Costa Rica', 'Nicaragua', 'Panamá', 'Rep. Dominicana'],
    'Asia': ['China', 'Corea del Sur', 'Filipinas', 'India', 'Japón', 'Tailandia'],
    'África': ['Angola', 'Cabo Verde', 'Sudáfrica'],
    'Oceanía': ['Nueva Zelanda'],
}

# Reverse: pais -> continente
PAIS_CONTINENTE = {}
for cont, paises in CONTINENTES.items():
    for p in paises:
        PAIS_CONTINENTE[p] = cont

# ── DB ────────────────────────────────────────────────────────────────────────

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS latas (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        marca            TEXT NOT NULL,
        modelo           TEXT,
        tipo_lata        TEXT,
        pais             TEXT,
        cantidad         INTEGER DEFAULT 1,
        foto_path        TEXT,
        notas            TEXT,
        fecha_adquisicion TEXT,
        fecha_carga      TEXT
    )''')

    # Migration: add columns if they don't exist yet
    existing = {row[1] for row in c.execute("PRAGMA table_info(latas)")}
    for col, defn in [
        ('foto_path',         'TEXT'),
        ('notas',             'TEXT'),
        ('fecha_adquisicion', 'TEXT'),
        ('sabor',             'INTEGER'),
    ]:
        if col not in existing:
            c.execute(f'ALTER TABLE latas ADD COLUMN {col} {defn}')

    # Índices para acelerar filtros y agrupaciones frecuentes
    for idx_sql in [
        'CREATE INDEX IF NOT EXISTS idx_latas_pais ON latas(pais)',
        'CREATE INDEX IF NOT EXISTS idx_latas_marca ON latas(marca)',
        'CREATE INDEX IF NOT EXISTS idx_latas_tipo ON latas(tipo_lata)',
        'CREATE INDEX IF NOT EXISTS idx_latas_sabor ON latas(sabor)',
    ]:
        c.execute(idx_sql)

    # Lista de deseos (latas que Pepe todavía no tiene pero quiere conseguir)
    c.execute('''CREATE TABLE IF NOT EXISTS wishlist (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        marca       TEXT NOT NULL,
        modelo      TEXT,
        pais        TEXT,
        prioridad   TEXT DEFAULT 'media',
        notas       TEXT,
        fecha_agregado TEXT
    )''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_wishlist_pais ON wishlist(pais)')

    # Buzón de latas anotadas desde la versión de internet, pendientes de que
    # la PC las traiga. No se borran al sincronizar, se marcan como tal: así
    # esta tabla queda como respaldo permanente de todo lo anotado afuera de casa.
    c.execute('''CREATE TABLE IF NOT EXISTS pendientes (
        id                INTEGER PRIMARY KEY AUTOINCREMENT,
        marca             TEXT NOT NULL,
        modelo            TEXT,
        tipo_lata         TEXT,
        pais              TEXT,
        cantidad          INTEGER DEFAULT 1,
        notas             TEXT,
        fecha_adquisicion TEXT,
        sabor             INTEGER,
        creado            TEXT,
        synced            INTEGER DEFAULT 0,
        synced_at         TEXT
    )''')

    conn.commit()

    c.execute('SELECT COUNT(*) FROM latas')
    if c.fetchone()[0] == 0:
        try:
            _import_excel(conn)
        except Exception:
            log.exception('Falló la importación inicial desde Excel')

    conn.close()

def _import_excel(conn):
    if not os.path.exists(EXCEL_PATH):
        return
    df = pd.read_excel(EXCEL_PATH, sheet_name='General', header=None)
    data = df.iloc[3:].copy()
    data.columns = ['marca', 'num', 'modelo', 'tipo_lata', 'pais', 'cantidad', 'subtotales']

    # Stop at the "Total" footer row
    total_rows = data[data['marca'].astype(str).str.strip() == 'Total'].index
    if len(total_rows):
        data = data.loc[:total_rows[0] - 1]

    SKIP_MARCAS = {'Marca', 'N° de países', 'Subtotales por país', 'nan'}
    SKIP_PAISES = {'Procedencia', 'N° de países', 'N° de marcas', 'nan'}

    c = conn.cursor()
    count = 0
    for _, row in data.iterrows():
        marca = str(row['marca']).strip() if pd.notna(row['marca']) else ''
        if not marca or marca in SKIP_MARCAS:
            continue
        pais = str(row['pais']).strip() if pd.notna(row['pais']) else ''
        if not pais or pais in SKIP_PAISES or pais.lstrip('-').isdigit():
            continue

        modelo = str(row['modelo']).strip() if pd.notna(row['modelo']) else None
        tipo   = str(row['tipo_lata']).strip() if pd.notna(row['tipo_lata']) else None
        if tipo   in ('nan', 'Tipo de lata', 'N° de países'): tipo   = None
        if modelo in ('nan',):                                  modelo = None

        try:
            cantidad = int(float(row['cantidad'])) if pd.notna(row['cantidad']) else 1
        except (ValueError, TypeError):
            cantidad = 1
        cantidad = max(cantidad, 1)

        c.execute(
            'INSERT INTO latas (marca, modelo, tipo_lata, pais, cantidad, fecha_carga) VALUES (?,?,?,?,?,?)',
            (marca, modelo, tipo, pais, cantidad, datetime.now().strftime('%Y-%m-%d'))
        )
        count += 1

    conn.commit()
    print(f'[import] {count} registros importados desde Excel')

# ── Acceso desde el celular (LAN) ───────────────────────────────────────────

def get_lan_ip():
    """IP de esta PC dentro de la red local (Wi-Fi de casa). No requiere internet:
    el 'connect' a 8.8.8.8 es UDP y solo sirve para que el SO elija la interfaz
    de salida, no llega a enviarse ningún paquete."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        return s.getsockname()[0]
    except OSError:
        return '127.0.0.1'
    finally:
        s.close()

# ── Acceso a la versión de internet (solo aplica en modo nube) ─────────────
# En la PC no hay login: es tu propia máquina. En la nube sí, porque el link
# queda expuesto a internet.

PUBLIC_PATHS = {'/login', '/logout', '/sw.js'}
TOKEN_PATHS  = {'/api/pendientes', '/api/pendientes/limpiar', '/api/push'}

@app.before_request
def _cloud_gate():
    if not CLOUD_MODE:
        return
    if request.path.startswith('/static/') or request.path in PUBLIC_PATHS:
        return
    if request.path in TOKEN_PATHS:
        token = request.args.get('token') or request.headers.get('X-Token')
        if token != CLOUD_PIN:
            return jsonify({'error': 'no autorizado'}), 401
        return
    if not session.get('ok'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'no autorizado'}), 401
        return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form.get('pin', '') == CLOUD_PIN:
            session['ok'] = True
            return redirect(url_for('index'))
        error = 'PIN incorrecto'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def _cloud_readonly():
    return jsonify({'error': 'Esta acción no está disponible en la versión de internet. Se hace desde la PC.'}), 403

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/sw.js')
def service_worker():
    """Servido desde la raíz (no /static/) para que el scope del service worker
    cubra toda la app y no solo /static/, que es lo que exigen los navegadores
    para permitir 'agregar a inicio' como PWA."""
    return send_file(os.path.join(BASE_DIR, 'static', 'sw.js'), mimetype='application/javascript')

@app.route('/api/lan-info')
def api_lan_info():
    return jsonify({'url': f'http://{get_lan_ip()}:{PORT}'})

@app.route('/api/qr')
def api_qr():
    url = f'http://{get_lan_ip()}:{PORT}'
    img = qrcode.make(url, box_size=8, border=2)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png')

@app.route('/api/stats')
def api_stats():
    conn = get_conn()
    c = conn.cursor()

    c.execute('SELECT COUNT(*) FROM latas');                              total_registros = c.fetchone()[0]
    c.execute('SELECT COALESCE(SUM(cantidad),0) FROM latas');             total_latas     = c.fetchone()[0]
    c.execute('SELECT COUNT(DISTINCT pais) FROM latas WHERE pais IS NOT NULL'); total_paises = c.fetchone()[0]
    c.execute('SELECT COUNT(DISTINCT marca) FROM latas');                 total_marcas    = c.fetchone()[0]

    # Por país (all)
    c.execute('SELECT pais, SUM(cantidad) as t FROM latas WHERE pais IS NOT NULL GROUP BY pais ORDER BY t DESC')
    todos_paises_raw = [(r['pais'], r['t']) for r in c.fetchall()]

    # Enrich with ISO2 and flag
    todos_paises = [
        {'pais': p, 'total': t, 'iso2': PAIS_ISO.get(p, ''), 'continente': PAIS_CONTINENTE.get(p, 'Otro')}
        for p, t in todos_paises_raw
    ]

    por_pais    = todos_paises[:15]
    map_data    = {}
    for p, t in todos_paises_raw:
        iso = PAIS_ISO.get(p)
        if iso:
            map_data[iso] = map_data.get(iso, 0) + t

    # Por tipo
    c.execute("SELECT COALESCE(tipo_lata,'Sin especificar') as tipo, SUM(cantidad) as t FROM latas GROUP BY tipo ORDER BY t DESC")
    por_tipo = [{'tipo': r['tipo'], 'total': r['t']} for r in c.fetchall()]

    # Top 15 marcas
    c.execute('SELECT marca, SUM(cantidad) as t FROM latas GROUP BY marca ORDER BY t DESC LIMIT 15')
    por_marca = [{'marca': r['marca'], 'total': r['t']} for r in c.fetchall()]

    # Continentes
    por_continente = []
    for cont, paises in CONTINENTES.items():
        ph = ','.join(['?'] * len(paises))
        c.execute(f'SELECT COALESCE(SUM(cantidad),0) FROM latas WHERE pais IN ({ph})', paises)
        por_continente.append({'continente': cont, 'total': c.fetchone()[0]})
    por_continente.sort(key=lambda x: -x['total'])

    # Timeline
    c.execute('''SELECT strftime('%Y-%m', fecha_adquisicion) as mes,
                        COUNT(*) as registros, SUM(cantidad) as total
                 FROM latas WHERE fecha_adquisicion IS NOT NULL
                 GROUP BY mes ORDER BY mes''')
    timeline = [{'mes': r['mes'], 'registros': r['registros'], 'total': r['total']} for r in c.fetchall()]

    # Sabor (escala hedónica 1-5)
    c.execute('SELECT AVG(sabor) FROM latas WHERE sabor IS NOT NULL')
    promedio_sabor = c.fetchone()[0]
    promedio_sabor = round(promedio_sabor, 2) if promedio_sabor is not None else None

    c.execute('SELECT sabor, COUNT(*) as n FROM latas WHERE sabor IS NOT NULL GROUP BY sabor ORDER BY sabor')
    conteo = {r['sabor']: r['n'] for r in c.fetchall()}
    por_sabor = [{'sabor': i, 'total': conteo.get(i, 0)} for i in range(1, 6)]

    c.execute('''SELECT marca, modelo, pais, sabor FROM latas
                 WHERE sabor IS NOT NULL ORDER BY sabor DESC, marca LIMIT 5''')
    top_sabor = [{'marca': r['marca'], 'modelo': r['modelo'], 'pais': r['pais'], 'sabor': r['sabor']} for r in c.fetchall()]

    # Treemap (top 40 marcas)
    c.execute('SELECT marca, SUM(cantidad) as t FROM latas GROUP BY marca ORDER BY t DESC LIMIT 40')
    treemap = [{'marca': r['marca'], 'total': r['t']} for r in c.fetchall()]

    # Wishlist: cuántas faltan y de qué países nuevos (que no están en la colección)
    c.execute('SELECT COUNT(*) FROM wishlist')
    wishlist_count = c.fetchone()[0]

    paises_tenidos = {p for p, _ in todos_paises_raw}
    c.execute('SELECT DISTINCT pais FROM wishlist WHERE pais IS NOT NULL')
    paises_wishlist_nuevos = sorted({r['pais'] for r in c.fetchall() if r['pais'] not in paises_tenidos})

    conn.close()
    return jsonify({
        'total_registros': total_registros,
        'total_latas':     total_latas,
        'total_paises':    total_paises,
        'total_marcas':    total_marcas,
        'por_pais':        por_pais,
        'todos_paises':    todos_paises,
        'map_data':        map_data,
        'por_tipo':        por_tipo,
        'por_marca':       por_marca,
        'por_continente':  por_continente,
        'timeline':        timeline,
        'treemap':         treemap,
        'promedio_sabor':  promedio_sabor,
        'por_sabor':       por_sabor,
        'top_sabor':       top_sabor,
        'wishlist_count':  wishlist_count,
        'paises_wishlist_nuevos': paises_wishlist_nuevos,
    })

@app.route('/api/latas')
def api_latas():
    q          = request.args.get('q', '').strip()
    pais       = request.args.get('pais', '').strip()
    tipo       = request.args.get('tipo', '').strip()
    continente = request.args.get('continente', '').strip()
    sabor_min  = request.args.get('sabor_min', '').strip()
    con_foto   = request.args.get('con_foto', '').strip()

    try:
        page = max(int(request.args.get('page', 1)), 1)
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('per_page', 50))
    except ValueError:
        per_page = 50
    per_page = max(1, min(per_page, MAX_PER_PAGE))

    conditions, params = [], []

    if q:
        conditions.append('(marca LIKE ? OR modelo LIKE ? OR notas LIKE ? OR pais LIKE ?)')
        params += [f'%{q}%'] * 4

    if pais:
        conditions.append('pais = ?')
        params.append(pais)
    elif continente and continente in CONTINENTES:
        ph = ','.join(['?'] * len(CONTINENTES[continente]))
        conditions.append(f'pais IN ({ph})')
        params += CONTINENTES[continente]

    if tipo:
        if tipo == 'Sin especificar':
            conditions.append('tipo_lata IS NULL')
        else:
            conditions.append('tipo_lata = ?')
            params.append(tipo)

    if sabor_min in ('1', '2', '3', '4', '5'):
        conditions.append('sabor >= ?')
        params.append(int(sabor_min))

    if con_foto in ('1', 'true'):
        conditions.append("foto_path IS NOT NULL AND foto_path != ''")

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    conn = get_conn()
    c = conn.cursor()

    c.execute(f'SELECT COUNT(*) FROM latas {where}', params)
    total = c.fetchone()[0]

    offset = (page - 1) * per_page
    c.execute(
        f'SELECT * FROM latas {where} ORDER BY pais, marca, modelo LIMIT ? OFFSET ?',
        params + [per_page, offset]
    )
    rows = []
    for r in c.fetchall():
        d = dict(r)
        d['iso2'] = PAIS_ISO.get(d.get('pais') or '', '')
        rows.append(d)

    c.execute('SELECT DISTINCT pais FROM latas WHERE pais IS NOT NULL ORDER BY pais')
    paises = [r['pais'] for r in c.fetchall()]

    c.execute('SELECT DISTINCT tipo_lata FROM latas WHERE tipo_lata IS NOT NULL ORDER BY tipo_lata')
    tipos = [r['tipo_lata'] for r in c.fetchall()]

    conn.close()
    return jsonify({'data': rows, 'total': total, 'paises': paises, 'tipos': tipos})

@app.route('/api/latas/<int:lid>')
def api_lata(lid):
    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT * FROM latas WHERE id=?', (lid,))
    row = c.fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'no encontrada'}), 404
    d = dict(row)
    d['iso2'] = PAIS_ISO.get(d.get('pais') or '', '')
    return jsonify(d)

def _parse_sabor(d):
    sabor = d.get('sabor')
    if sabor in (None, '', 0):
        return None
    try:
        sabor = int(sabor)
    except (ValueError, TypeError):
        return None
    return sabor if 1 <= sabor <= 5 else None

def _validate_lata(d):
    """Devuelve un mensaje de error si los datos no son válidos, o None si está OK."""
    if not isinstance(d, dict):
        return 'Datos inválidos'
    if not (d.get('marca') or '').strip():
        return 'La marca es obligatoria'
    if not (d.get('pais') or '').strip():
        return 'El país es obligatorio'
    try:
        if int(d.get('cantidad', 1)) < 1:
            return 'La cantidad debe ser al menos 1'
    except (ValueError, TypeError):
        return 'La cantidad debe ser un número'
    sabor = d.get('sabor')
    if sabor not in (None, '', 0):
        try:
            if not (1 <= int(sabor) <= 5):
                return 'El sabor debe estar entre 1 y 5'
        except (ValueError, TypeError):
            return 'El sabor debe ser un número entre 1 y 5'
    return None

def _insert_pendiente(d):
    conn = get_conn()
    conn.execute(
        '''INSERT INTO pendientes (marca, modelo, tipo_lata, pais, cantidad, notas, fecha_adquisicion, sabor, creado)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (
            (d.get('marca') or '').strip(),
            (d.get('modelo') or '').strip() or None,
            d.get('tipo_lata') or None,
            (d.get('pais') or '').strip() or None,
            max(int(d.get('cantidad', 1)), 1),
            (d.get('notas') or '').strip() or None,
            d.get('fecha_adquisicion') or None,
            _parse_sabor(d),
            datetime.now().strftime('%Y-%m-%d %H:%M'),
        )
    )
    conn.commit()
    new_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.close()
    return new_id

@app.route('/api/latas', methods=['POST'])
def api_create():
    d = request.get_json(silent=True) or {}
    err = _validate_lata(d)
    if err:
        return jsonify({'error': err}), 400

    if CLOUD_MODE:
        new_id = _insert_pendiente(d)
        return jsonify({'id': new_id, 'ok': True, 'pendiente': True})

    marca  = (d.get('marca') or '').strip()
    modelo = (d.get('modelo') or '').strip() or None
    pais   = (d.get('pais') or '').strip() or None

    if not d.get('force'):
        dupconn = get_conn()
        dupc = dupconn.cursor()
        dupc.execute(
            '''SELECT id, cantidad FROM latas
               WHERE LOWER(marca) = LOWER(?)
                 AND LOWER(COALESCE(modelo,'')) = LOWER(COALESCE(?,''))
                 AND LOWER(COALESCE(pais,''))   = LOWER(COALESCE(?,''))''',
            (marca, modelo, pais)
        )
        dup = dupc.fetchone()
        dupconn.close()
        if dup:
            return jsonify({
                'duplicate': True,
                'existing_id': dup['id'],
                'existing_cantidad': dup['cantidad'],
            }), 409

    conn = get_conn()
    c = conn.cursor()
    c.execute(
        '''INSERT INTO latas (marca, modelo, tipo_lata, pais, cantidad, notas, fecha_adquisicion, sabor, fecha_carga)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (
            (d.get('marca') or '').strip(),
            (d.get('modelo') or '').strip() or None,
            d.get('tipo_lata') or None,
            (d.get('pais') or '').strip() or None,
            max(int(d.get('cantidad', 1)), 1),
            (d.get('notas') or '').strip() or None,
            d.get('fecha_adquisicion') or None,
            _parse_sabor(d),
            datetime.now().strftime('%Y-%m-%d'),
        )
    )
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/latas/<int:lid>', methods=['PUT'])
def api_update(lid):
    if CLOUD_MODE:
        return _cloud_readonly()
    d = request.get_json(silent=True) or {}
    err = _validate_lata(d)
    if err:
        return jsonify({'error': err}), 400

    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM latas WHERE id=?', (lid,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404

    conn.execute(
        '''UPDATE latas SET marca=?, modelo=?, tipo_lata=?, pais=?,
           cantidad=?, notas=?, fecha_adquisicion=?, sabor=? WHERE id=?''',
        (
            (d.get('marca') or '').strip(),
            (d.get('modelo') or '').strip() or None,
            d.get('tipo_lata') or None,
            (d.get('pais') or '').strip() or None,
            max(int(d.get('cantidad', 1)), 1),
            (d.get('notas') or '').strip() or None,
            d.get('fecha_adquisicion') or None,
            _parse_sabor(d),
            lid,
        )
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/latas/<int:lid>/sabor', methods=['PATCH'])
def api_lata_sabor(lid):
    """Calificación rápida (estrellas clickeables desde la tabla, sin abrir el modal)."""
    if CLOUD_MODE:
        return _cloud_readonly()
    d = request.get_json(silent=True) or {}
    if d.get('sabor') not in (None, '', 0):
        try:
            if not (1 <= int(d['sabor']) <= 5):
                return jsonify({'error': 'El sabor debe estar entre 1 y 5'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'El sabor debe ser un número entre 1 y 5'}), 400
    sabor = _parse_sabor(d)

    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM latas WHERE id=?', (lid,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    conn.execute('UPDATE latas SET sabor=? WHERE id=?', (sabor, lid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'sabor': sabor})

@app.route('/api/latas/<int:lid>', methods=['DELETE'])
def api_delete(lid):
    if CLOUD_MODE:
        return _cloud_readonly()
    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM latas WHERE id=?', (lid,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    conn.execute('DELETE FROM latas WHERE id=?', (lid,))
    conn.commit()
    conn.close()
    # Remove photo if exists
    for ext in FOTO_EXTS:
        p = os.path.join(FOTOS_DIR, f'{lid}.{ext}')
        if os.path.exists(p):
            os.remove(p)
    return jsonify({'ok': True})

# ── Wishlist (latas que Pepe quiere conseguir) ─────────────────────────────────

PRIORIDADES_VALIDAS = {'alta', 'media', 'baja'}

def _validate_wishlist(d):
    if not isinstance(d, dict):
        return 'Datos inválidos'
    if not (d.get('marca') or '').strip():
        return 'La marca es obligatoria'
    prioridad = d.get('prioridad') or 'media'
    if prioridad not in PRIORIDADES_VALIDAS:
        return 'La prioridad debe ser alta, media o baja'
    return None

@app.route('/api/wishlist')
def api_wishlist_list():
    conn = get_conn()
    c = conn.cursor()
    c.execute('''SELECT * FROM wishlist ORDER BY
                 CASE prioridad WHEN 'alta' THEN 0 WHEN 'media' THEN 1 ELSE 2 END,
                 marca, modelo''')
    rows = []
    for r in c.fetchall():
        d = dict(r)
        d['iso2'] = PAIS_ISO.get(d.get('pais') or '', '')
        rows.append(d)
    conn.close()
    return jsonify({'data': rows})

@app.route('/api/wishlist', methods=['POST'])
def api_wishlist_create():
    d = request.get_json(silent=True) or {}
    err = _validate_wishlist(d)
    if err:
        return jsonify({'error': err}), 400

    conn = get_conn()
    c = conn.cursor()
    c.execute(
        '''INSERT INTO wishlist (marca, modelo, pais, prioridad, notas, fecha_agregado)
           VALUES (?,?,?,?,?,?)''',
        (
            (d.get('marca') or '').strip(),
            (d.get('modelo') or '').strip() or None,
            (d.get('pais') or '').strip() or None,
            d.get('prioridad') or 'media',
            (d.get('notas') or '').strip() or None,
            datetime.now().strftime('%Y-%m-%d'),
        )
    )
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/wishlist/<int:wid>', methods=['PUT'])
def api_wishlist_update(wid):
    if CLOUD_MODE:
        return _cloud_readonly()
    d = request.get_json(silent=True) or {}
    err = _validate_wishlist(d)
    if err:
        return jsonify({'error': err}), 400

    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM wishlist WHERE id=?', (wid,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404

    conn.execute(
        '''UPDATE wishlist SET marca=?, modelo=?, pais=?, prioridad=?, notas=? WHERE id=?''',
        (
            (d.get('marca') or '').strip(),
            (d.get('modelo') or '').strip() or None,
            (d.get('pais') or '').strip() or None,
            d.get('prioridad') or 'media',
            (d.get('notas') or '').strip() or None,
            wid,
        )
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/wishlist/<int:wid>', methods=['DELETE'])
def api_wishlist_delete(wid):
    if CLOUD_MODE:
        return _cloud_readonly()
    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM wishlist WHERE id=?', (wid,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    conn.execute('DELETE FROM wishlist WHERE id=?', (wid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/wishlist/<int:wid>/conseguir', methods=['POST'])
def api_wishlist_conseguir(wid):
    """Pasa un ítem de la wishlist a la colección real."""
    if CLOUD_MODE:
        return _cloud_readonly()
    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT * FROM wishlist WHERE id=?', (wid,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404

    today = datetime.now().strftime('%Y-%m-%d')
    c.execute(
        '''INSERT INTO latas (marca, modelo, tipo_lata, pais, cantidad, notas, fecha_adquisicion, fecha_carga)
           VALUES (?,?,?,?,?,?,?,?)''',
        (row['marca'], row['modelo'], None, row['pais'], 1, row['notas'], today, today)
    )
    new_id = c.lastrowid
    conn.execute('DELETE FROM wishlist WHERE id=?', (wid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': new_id})

# ── Sincronización con el "buzón" en la nube (cloud/app.py) ────────────────
# Ese buzón vive en un hosting gratuito 24/7 y solo junta latas anotadas
# desde el celular estando afuera de casa. Acá las traemos, las mergeamos
# con la colección real (sumando cantidad si ya existía) y limpiamos el
# buzón para la próxima.

def _load_sync_config():
    if not os.path.exists(SYNC_CONFIG_PATH):
        return None
    with open(SYNC_CONFIG_PATH, encoding='utf-8') as f:
        return json.load(f)

def _importar_pendiente(conn, d):
    marca  = (d.get('marca') or '').strip()
    if not marca:
        return
    modelo = (d.get('modelo') or '').strip() or None
    pais   = (d.get('pais') or '').strip() or None
    cantidad = max(int(d.get('cantidad') or 1), 1)

    c = conn.cursor()
    c.execute(
        '''SELECT id, cantidad FROM latas
           WHERE LOWER(marca) = LOWER(?)
             AND LOWER(COALESCE(modelo,'')) = LOWER(COALESCE(?,''))
             AND LOWER(COALESCE(pais,''))   = LOWER(COALESCE(?,''))''',
        (marca, modelo, pais)
    )
    dup = c.fetchone()
    if dup:
        conn.execute('UPDATE latas SET cantidad=? WHERE id=?', (dup['cantidad'] + cantidad, dup['id']))
    else:
        conn.execute(
            '''INSERT INTO latas (marca, modelo, tipo_lata, pais, cantidad, notas, fecha_adquisicion, sabor, fecha_carga)
               VALUES (?,?,?,?,?,?,?,?,?)''',
            (
                marca, modelo, d.get('tipo_lata') or None, pais, cantidad,
                d.get('notas') or None, d.get('fecha_adquisicion') or None, d.get('sabor'),
                datetime.now().strftime('%Y-%m-%d'),
            )
        )

@app.route('/api/sync', methods=['POST'])
def api_sync():
    cfg = _load_sync_config()
    if not cfg or not cfg.get('cloud_url') or not cfg.get('token'):
        return jsonify({'error': f'Falta configurar la sincronización. Creá el archivo {SYNC_CONFIG_PATH} (ver sync_config.example.json en el proyecto como modelo)'}), 400

    base = cfg['cloud_url'].rstrip('/')
    try:
        with urllib.request.urlopen(f"{base}/api/pendientes?token={cfg['token']}", timeout=10) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
    except (urllib.error.URLError, TimeoutError, ValueError) as e:
        return jsonify({'error': f'No se pudo conectar con el buzón en la nube: {e}'}), 502

    pendientes = payload.get('data', [])
    if not pendientes:
        return jsonify({'ok': True, 'importados': 0, 'detalle': []})

    conn = get_conn()
    for p in pendientes:
        _importar_pendiente(conn, p)
    conn.commit()
    conn.close()

    ids = [p['id'] for p in pendientes]
    try:
        req = urllib.request.Request(
            f'{base}/api/pendientes/limpiar',
            data=json.dumps({'ids': ids}).encode('utf-8'),
            headers={'Content-Type': 'application/json', 'X-Token': cfg['token']},
            method='POST',
        )
        urllib.request.urlopen(req, timeout=10)
    except (urllib.error.URLError, TimeoutError):
        log.exception('Se importaron las latas del buzón pero no se pudo limpiarlo en la nube')

    detalle = [f"{p['marca']} ({p.get('pais') or 'sin país'})" for p in pendientes]
    return jsonify({'ok': True, 'importados': len(pendientes), 'detalle': detalle})

# ── Lado nube: recibe pull de pendientes y push de la colección completa ───
# (solo se usan cuando este mismo app.py corre en PythonAnywhere en CLOUD_MODE)

@app.route('/api/pendientes')
def api_pendientes():
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        'SELECT * FROM pendientes WHERE synced=0 OR synced IS NULL ORDER BY id'
    ).fetchall()]
    conn.close()
    return jsonify({'data': rows})

@app.route('/api/pendientes/limpiar', methods=['POST'])
def api_pendientes_limpiar():
    """Marca como sincronizado en vez de borrar: queda como respaldo permanente
    de todo lo anotado desde la versión de internet."""
    ids = (request.get_json(silent=True) or {}).get('ids', [])
    if not ids:
        return jsonify({'error': 'no ids'}), 400
    conn = get_conn()
    ph = ','.join('?' * len(ids))
    conn.execute(
        f'UPDATE pendientes SET synced=1, synced_at=? WHERE id IN ({ph})',
        [datetime.now().strftime('%Y-%m-%d %H:%M')] + ids
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

FOTO_FNAME_RE = re.compile(r'^\d+\.(jpg|jpeg|png|gif|webp)$')

@app.route('/api/push', methods=['POST'])
def api_push():
    """Recibe desde la PC una foto completa de latas+wishlist (y las fotos
    comprimidas que corresponden) y reemplaza lo que hay en la nube. No toca
    la tabla 'pendientes': lo anotado ahí desde el celular/internet que
    todavía no viajó a la PC se conserva intacto."""
    payload = request.form.get('data')
    if not payload:
        return jsonify({'error': 'sin datos'}), 400
    try:
        data = json.loads(payload)
    except ValueError:
        return jsonify({'error': 'JSON inválido'}), 400

    conn = get_conn()
    conn.execute('DELETE FROM latas')
    for l in data.get('latas', []):
        conn.execute(
            '''INSERT INTO latas (id, marca, modelo, tipo_lata, pais, cantidad, foto_path, notas, fecha_adquisicion, fecha_carga, sabor)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
            (l.get('id'), l.get('marca'), l.get('modelo'), l.get('tipo_lata'), l.get('pais'),
             l.get('cantidad', 1), l.get('foto_path'), l.get('notas'), l.get('fecha_adquisicion'),
             l.get('fecha_carga'), l.get('sabor'))
        )
    conn.execute('DELETE FROM wishlist')
    for w in data.get('wishlist', []):
        conn.execute(
            '''INSERT INTO wishlist (id, marca, modelo, pais, prioridad, notas, fecha_agregado)
               VALUES (?,?,?,?,?,?,?)''',
            (w.get('id'), w.get('marca'), w.get('modelo'), w.get('pais'),
             w.get('prioridad', 'media'), w.get('notas'), w.get('fecha_agregado'))
        )
    conn.commit()
    conn.close()

    os.makedirs(FOTOS_DIR, exist_ok=True)
    fotos_guardadas = 0
    for fname, f in request.files.items():
        if FOTO_FNAME_RE.match(fname):
            f.save(os.path.join(FOTOS_DIR, fname))
            fotos_guardadas += 1

    return jsonify({
        'ok': True,
        'latas': len(data.get('latas', [])),
        'wishlist': len(data.get('wishlist', [])),
        'fotos': fotos_guardadas,
    })

# ── Lado PC: arma y manda el push de arriba ─────────────────────────────────

@app.route('/api/push-cloud', methods=['POST'])
def api_push_cloud():
    if CLOUD_MODE:
        return _cloud_readonly()
    cfg = _load_sync_config()
    if not cfg or not cfg.get('cloud_url') or not cfg.get('token'):
        return jsonify({'error': f'Falta configurar la sincronización. Creá el archivo {SYNC_CONFIG_PATH} (ver sync_config.example.json en el proyecto como modelo)'}), 400

    conn = get_conn()
    latas    = [dict(r) for r in conn.execute('SELECT * FROM latas').fetchall()]
    wishlist = [dict(r) for r in conn.execute('SELECT * FROM wishlist').fetchall()]
    conn.close()

    # Comprimimos cada foto a un tamaño chico antes de mandarla, para que la
    # versión de internet sea liviana y rápida aunque haya cientos de fotos.
    boundary = '----latasboundary' + secrets.token_hex(8)
    photo_parts = []
    for l in latas:
        fp = l.get('foto_path')
        if not fp:
            continue
        local_path = os.path.join(BASE_DIR, fp.lstrip('/'))
        if not os.path.exists(local_path):
            l['foto_path'] = None
            continue
        try:
            with Image.open(local_path) as img:
                img = img.convert('RGB')
                img.thumbnail((400, 400))
                buf = io.BytesIO()
                img.save(buf, format='JPEG', quality=70)
        except (UnidentifiedImageError, OSError):
            l['foto_path'] = None
            continue
        fname = f"{l['id']}.jpg"
        l['foto_path'] = f'/static/fotos/{fname}'
        photo_parts.append((fname, buf.getvalue()))

    payload = json.dumps({'latas': latas, 'wishlist': wishlist}).encode('utf-8')

    parts = [
        f'--{boundary}\r\nContent-Disposition: form-data; name="data"\r\n\r\n'.encode('utf-8') + payload + b'\r\n'
    ]
    for fname, data_bytes in photo_parts:
        parts.append(
            (f'--{boundary}\r\nContent-Disposition: form-data; name="{fname}"; filename="{fname}"\r\n'
             f'Content-Type: image/jpeg\r\n\r\n').encode('utf-8') + data_bytes + b'\r\n'
        )
    parts.append(f'--{boundary}--\r\n'.encode('utf-8'))
    body = b''.join(parts)

    base = cfg['cloud_url'].rstrip('/')
    req = urllib.request.Request(
        f"{base}/api/push?token={cfg['token']}",
        data=body,
        headers={'Content-Type': f'multipart/form-data; boundary={boundary}'},
        method='POST',
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode('utf-8'))
    except (urllib.error.URLError, TimeoutError, ValueError) as e:
        return jsonify({'error': f'No se pudo conectar con la app en la nube: {e}'}), 502

    return jsonify({'ok': True, **result})

@app.route('/api/latas/<int:lid>/foto', methods=['POST'])
def api_foto(lid):
    if CLOUD_MODE:
        return _cloud_readonly()
    if 'foto' not in request.files:
        return jsonify({'error': 'no file'}), 400
    f = request.files['foto']
    if not f.filename:
        return jsonify({'error': 'empty'}), 400
    if '.' not in f.filename:
        return jsonify({'error': 'formato no soportado'}), 400
    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in FOTO_EXTS:
        return jsonify({'error': 'formato no soportado'}), 400

    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT id FROM latas WHERE id=?', (lid,))
    if not c.fetchone():
        conn.close()
        return jsonify({'error': 'no encontrada'}), 404
    conn.close()

    os.makedirs(FOTOS_DIR, exist_ok=True)
    tmp_path = os.path.join(FOTOS_DIR, f'_tmp_{lid}.{ext}')
    f.save(tmp_path)

    # Verificamos que el archivo sea realmente una imagen (no solo que tenga
    # extensión de imagen) antes de aceptarlo.
    try:
        with Image.open(tmp_path) as img:
            img.verify()
    except (UnidentifiedImageError, OSError):
        os.remove(tmp_path)
        return jsonify({'error': 'el archivo no es una imagen válida'}), 400

    # Remove old photos for this id
    for old_ext in FOTO_EXTS:
        old = os.path.join(FOTOS_DIR, f'{lid}.{old_ext}')
        if os.path.exists(old):
            os.remove(old)

    filename  = f'{lid}.{ext}'
    save_path = os.path.join(FOTOS_DIR, filename)
    shutil.move(tmp_path, save_path)

    foto_url = f'/static/fotos/{filename}'
    conn = get_conn()
    conn.execute('UPDATE latas SET foto_path=? WHERE id=?', (foto_url, lid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'foto_url': foto_url})

@app.route('/api/opciones')
def api_opciones():
    conn = get_conn()
    c = conn.cursor()
    c.execute('SELECT DISTINCT pais FROM latas WHERE pais IS NOT NULL ORDER BY pais')
    paises = [{'pais': r['pais'], 'iso2': PAIS_ISO.get(r['pais'], '')} for r in c.fetchall()]
    c.execute('SELECT DISTINCT tipo_lata FROM latas WHERE tipo_lata IS NOT NULL ORDER BY tipo_lata')
    tipos = [r['tipo_lata'] for r in c.fetchall()]
    c.execute('SELECT DISTINCT marca FROM latas ORDER BY marca')
    marcas = [r['marca'] for r in c.fetchall()]
    conn.close()
    continentes = list(CONTINENTES.keys())
    # Catálogo completo de países conocidos (para el <select> del formulario, no solo
    # los que ya están en la colección) — mantiene las estadísticas por país consistentes.
    paises_catalogo = [{'pais': p, 'iso2': PAIS_ISO[p]} for p in sorted(PAIS_ISO.keys())]
    return jsonify({
        'paises': paises, 'tipos': tipos, 'marcas': marcas, 'continentes': continentes,
        'paises_catalogo': paises_catalogo,
    })

@app.route('/api/exportar')
def api_exportar():
    conn = get_conn()
    df = pd.read_sql_query(
        'SELECT marca, modelo, tipo_lata, pais, cantidad, sabor, notas, fecha_adquisicion FROM latas ORDER BY pais, marca, modelo',
        conn
    )
    conn.close()

    df.columns = ['Marca', 'Modelo', 'Tipo de lata', 'Procedencia', 'Cantidad', 'Sabor (1-5)', 'Notas', 'Fecha de adquisición']
    df = df.fillna('')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Colección'

    ws.merge_cells('A1:H1')
    tc = ws['A1']
    tc.value = '🍺 Colección de Latas de Cerveza – Pepe'
    tc.font  = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    tc.fill  = PatternFill('solid', start_color='1A3C5E')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = ['Marca', 'Modelo', 'Tipo de lata', 'Procedencia', 'Cantidad', 'Sabor (1-5)', 'Notas', 'Fecha de adquisición']
    hfill = PatternFill('solid', start_color='2E6DA4')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt    = PatternFill('solid', start_color='EEF4FB')
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if ri % 2 == 0:
                cell.fill = alt

    for i, w in enumerate([25, 20, 15, 20, 10, 12, 30, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Resumen sheet
    ws2 = wb.create_sheet('Resumen')
    conn2 = get_conn()
    c2 = conn2.cursor()
    ws2['A1'] = '📊 Resumen'
    ws2['A1'].font = Font(name='Arial', size=13, bold=True, color='1A3C5E')
    ws2['A2'] = f'Exportado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws2['A2'].font = Font(name='Arial', size=9, color='888888')

    c2.execute('SELECT COUNT(*) FROM latas');                   ws2['A4'] = 'Modelos distintos'; ws2['B4'] = c2.fetchone()[0]
    c2.execute('SELECT COALESCE(SUM(cantidad),0) FROM latas');  ws2['A5'] = 'Total latas';        ws2['B5'] = c2.fetchone()[0]
    c2.execute('SELECT COUNT(DISTINCT pais) FROM latas WHERE pais IS NOT NULL'); ws2['A6'] = 'Países'; ws2['B6'] = c2.fetchone()[0]
    c2.execute('SELECT COUNT(DISTINCT marca) FROM latas');       ws2['A7'] = 'Marcas';             ws2['B7'] = c2.fetchone()[0]

    ws2['A9'] = 'Latas por país'
    ws2['A9'].font = Font(bold=True)
    c2.execute('SELECT pais, SUM(cantidad) FROM latas WHERE pais IS NOT NULL GROUP BY pais ORDER BY 2 DESC')
    for i, (p, t) in enumerate(c2.fetchall(), 10):
        ws2.cell(row=i, column=1, value=p)
        ws2.cell(row=i, column=2, value=t)
    conn2.close()

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'coleccion_latas_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/wishlist/exportar')
def api_wishlist_exportar():
    """Exporta la lista de deseos a un Excel prolijo, pensado para compartir
    con alguien que quiera regalarle una lata a Pepe."""
    conn = get_conn()
    df = pd.read_sql_query(
        '''SELECT marca, modelo, pais, prioridad, notas FROM wishlist
           ORDER BY CASE prioridad WHEN 'alta' THEN 0 WHEN 'media' THEN 1 ELSE 2 END, marca, modelo''',
        conn
    )
    conn.close()

    df.columns = ['Marca', 'Modelo', 'País', 'Prioridad', 'Notas']
    df['Prioridad'] = df['Prioridad'].map({'alta': 'Alta', 'media': 'Media', 'baja': 'Baja'}).fillna(df['Prioridad'])
    df = df.fillna('')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Lista de deseos'

    ws.merge_cells('A1:E1')
    tc = ws['A1']
    tc.value = '💗 Latas que Pepe quiere conseguir'
    tc.font  = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    tc.fill  = PatternFill('solid', start_color='1A3C5E')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    headers = ['Marca', 'Modelo', 'País', 'Prioridad', 'Notas']
    hfill = PatternFill('solid', start_color='2E6DA4')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt    = PatternFill('solid', start_color='EEF4FB')
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if ri % 2 == 0:
                cell.fill = alt

    for i, w in enumerate([22, 20, 18, 12, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f'lista_de_deseos_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/backup')
def api_backup():
    """Descarga una copia de la base de datos actual (.db) como respaldo."""
    if not os.path.exists(DB_PATH):
        return jsonify({'error': 'base de datos no encontrada'}), 404
    fname = f'latas_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
    return send_file(DB_PATH, as_attachment=True, download_name=fname,
                      mimetype='application/octet-stream')

@app.route('/api/restore', methods=['POST'])
def api_restore():
    """Restaura la base de datos desde un archivo .db subido por el usuario
    (por ejemplo, uno descargado antes con /api/backup). Antes de sobreescribir,
    guarda una copia de seguridad del estado actual por las dudas."""
    if CLOUD_MODE:
        return _cloud_readonly()
    if 'backup' not in request.files:
        return jsonify({'error': 'no se envió ningún archivo'}), 400
    f = request.files['backup']
    if not f.filename or not f.filename.lower().endswith('.db'):
        return jsonify({'error': 'el archivo debe ser una copia de seguridad .db'}), 400

    os.makedirs(BACKUP_DIR, exist_ok=True)
    tmp_path = os.path.join(BACKUP_DIR, '_restore_tmp.db')
    f.save(tmp_path)

    test_conn = sqlite3.connect(tmp_path)
    try:
        tc = test_conn.cursor()
        tc.execute('PRAGMA quick_check')
        ok = tc.fetchone()[0] == 'ok'
        tc.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='latas'")
        has_table = tc.fetchone() is not None
    except sqlite3.DatabaseError:
        ok, has_table = False, False
    finally:
        test_conn.close()

    if not ok or not has_table:
        os.remove(tmp_path)
        return jsonify({'error': 'el archivo no es una copia de seguridad válida'}), 400

    if os.path.exists(DB_PATH):
        safety = os.path.join(BACKUP_DIR, f'antes_de_restaurar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db')
        shutil.copy2(DB_PATH, safety)

    shutil.move(tmp_path, DB_PATH)
    return jsonify({'ok': True})

def _auto_backup():
    """Guarda una copia diaria de la base de datos en /backups (máx. 14)."""
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        today = datetime.now().strftime('%Y%m%d')
        dest = os.path.join(BACKUP_DIR, f'latas_{today}.db')
        if not os.path.exists(dest) and os.path.exists(DB_PATH):
            shutil.copy2(DB_PATH, dest)
        backups = sorted(
            (f for f in os.listdir(BACKUP_DIR) if f.startswith('latas_') and f.endswith('.db')),
        )
        for old in backups[:-14]:
            os.remove(os.path.join(BACKUP_DIR, old))
    except Exception:
        log.exception('No se pudo generar el backup automático')

# ── Manejo de errores ──────────────────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'recurso no encontrado'}), 404
    return render_template('index.html'), 404

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'el archivo es demasiado grande (máx. 12 MB)'}), 413

@app.errorhandler(500)
def server_error(e):
    log.exception('Error interno')
    if request.path.startswith('/api/'):
        return jsonify({'error': 'error interno del servidor'}), 500
    return render_template('index.html'), 500

os.makedirs(FOTOS_DIR, exist_ok=True)
init_db()

if __name__ == '__main__':
    _auto_backup()
    lan_ip = get_lan_ip()
    print('\n🍺  Colección de Latas de Pepe arrancando...')
    print(f'   En esta PC:       http://localhost:{PORT}')
    print(f'   Desde el celular: http://{lan_ip}:{PORT}  (misma red Wi-Fi, sin internet)\n')
    app.run(debug=False, host='0.0.0.0', port=PORT)
